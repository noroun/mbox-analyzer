"""
Tests de non-régression sur la détection.

Construit un MBOX synthétique avec des cas adverses connus et vérifie que la
classification reste correcte après les améliorations.
"""

from __future__ import annotations

import mailbox
import sys
from email.message import EmailMessage
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from mbox_analyzer import (  # noqa: E402
    SERVICE_DOMAINS,
    is_automated_sender,
    normalize_domain,
    analyse,
)


# --------- Tests unitaires ---------


class TestServiceDomainsExactMatch:
    def test_apple_dot_com_is_service(self):
        assert "apple.com" in SERVICE_DOMAINS

    def test_snapple_is_not_service(self):
        # Le bug historique : "apple" en substring matchait snapple.com
        assert "snapple.com" not in SERVICE_DOMAINS

    def test_googleberry_not_service(self):
        assert "googleberry.fr" not in SERVICE_DOMAINS

    def test_thebaystreet_not_ebay(self):
        assert "thebaystreet.com" not in SERVICE_DOMAINS


class TestIsAutomatedSender:
    @pytest.mark.parametrize("addr", [
        "noreply@x.com",
        "no-reply@y.fr",
        "donotreply@bank.com",
        "do-not-reply@svc.io",
        "notifications@github.com",
        "notification@linkedin.com",
        "alerts@grafana.io",
        "info@startup.fr",
        "support@stripe.com",
        "service-client@sncf.fr",
    ])
    def test_automated(self, addr):
        assert is_automated_sender(addr), f"{addr} devrait être automatisé"

    @pytest.mark.parametrize("addr", [
        "alice@gmail.com",
        "jean.dupont@perso.fr",
        "norou@example.com",
        "marie-claire@asso.org",
    ])
    def test_human(self, addr):
        assert not is_automated_sender(addr), f"{addr} ne devrait pas être automatisé"

    def test_invalid(self):
        assert not is_automated_sender("")
        assert not is_automated_sender("bogus")


class TestNormalizeDomain:
    @pytest.mark.parametrize("raw,expected", [
        ("accounts.google.com", "google.com"),
        ("mail.notion.so", "notion.so"),
        ("foo.bar.co.uk", "bar.co.uk"),
        ("example.com", "example.com"),
    ])
    def test_normalize(self, raw, expected):
        assert normalize_domain(raw) == expected


# --------- Test end-to-end avec MBOX synthétique ---------


USER_EMAIL = "norou@example.com"


def _msg(*, sender, subject, to=USER_EMAIL, body="hello", labels="Inbox",
         date="Mon, 01 Jan 2024 12:00:00 +0000",
         list_unsubscribe=None, precedence=None, auto_submitted=None):
    m = EmailMessage()
    m["From"] = sender
    m["To"] = to
    m["Delivered-To"] = USER_EMAIL
    m["Subject"] = subject
    m["Date"] = date
    m["X-Gmail-Labels"] = labels
    if list_unsubscribe:
        m["List-Unsubscribe"] = list_unsubscribe
    if precedence:
        m["Precedence"] = precedence
    if auto_submitted:
        m["Auto-Submitted"] = auto_submitted
    m.set_content(body)
    return m


@pytest.fixture
def synthetic_mbox(tmp_path):
    """Construit un MBOX adversarial couvrant les cas de régression connus."""
    mbox_path = tmp_path / "test.mbox"
    box = mailbox.mbox(str(mbox_path))
    box.lock()

    # Suffisamment de mails pour que le user_emails seuil (>=10) soit atteint
    # → on duplique chaque scénario assez pour passer les 10 occurrences de To.
    # En pratique, on en met >= 10 au total mais avec un mélange de scénarios.

    # 1) Amazon : 3 confirmations de commande avec List-Unsubscribe + sujet transactionnel
    #    → doit être COMPTE (transaction), PAS newsletter.
    for i, ref in enumerate(["12345", "67890", "55555"]):
        box.add(_msg(
            sender=f"auto-confirm@amazon.fr",
            subject=f"Votre commande #{ref}",
            list_unsubscribe="<https://amazon.fr/unsub>",
            date=f"Mon, 0{i+1} Jan 2024 10:00:00 +0000",
        ))

    # 2) Substack : 3 newsletters sans Precedence/Auto-Submitted
    #    → doit être NEWSLETTER (corroboré par count >= 2 + non-transactionnel).
    for i in range(3):
        box.add(_msg(
            sender="weekly@substack.com",
            subject=f"Weekly digest #{i+1}",
            list_unsubscribe="<https://substack.com/unsub?token=abc>",
            date=f"Sun, 0{i+1} Feb 2024 09:00:00 +0000",
        ))

    # 3) Mailchimp newsletter avec Precedence: bulk dès le 1er mail
    #    → newsletter dès le premier (signal fort).
    box.add(_msg(
        sender="news@mailchimp-customer.fr",
        subject="Hello from us",
        list_unsubscribe="<https://mailchimp.com/unsub>",
        precedence="bulk",
        date="Wed, 03 Mar 2024 08:00:00 +0000",
    ))

    # 4) Conversation humaine : 4 mails depuis ami@perso.fr SANS List-Unsubscribe
    #    et expéditeur non-automatisé → ne doit PAS devenir compte.
    for i in range(4):
        box.add(_msg(
            sender="ami@perso.fr",
            subject=f"Re: déjeuner samedi {i}",
            date=f"Thu, 0{i+1} Apr 2024 14:00:00 +0000",
        ))

    # 5) Domaine "snapple.com" (faux positif "apple") — sender automatisé,
    #    mais le service-known doit rester False.
    box.add(_msg(
        sender="info@snapple.com",
        subject="Notre nouveau goût",
        list_unsubscribe="<https://snapple.com/unsub>",
        precedence="bulk",
        date="Fri, 05 May 2024 10:00:00 +0000",
    ))

    # 6) 2FA Google → doit être compte (auth).
    box.add(_msg(
        sender="no-reply@accounts.google.com",
        subject="Code de vérification 482931",
        date="Sat, 06 Jun 2024 11:00:00 +0000",
    ))

    # 7) LinkedIn (service connu, normalisé en linkedin.com)
    box.add(_msg(
        sender="messaging-digest-noreply@linkedin.com",
        subject="Vous avez 3 nouvelles notifications",
        date="Sun, 07 Jul 2024 12:00:00 +0000",
    ))

    # 8) Inscription explicite
    box.add(_msg(
        sender="welcome@notion.so",
        subject="Bienvenue sur Notion",
        date="Mon, 08 Aug 2024 13:00:00 +0000",
    ))

    box.flush()
    box.unlock()
    box.close()
    return mbox_path


def test_end_to_end(synthetic_mbox, tmp_path):
    out_dir = tmp_path / "out"
    progress = []
    analyse(str(synthetic_mbox), str(out_dir), lambda i, t: progress.append((i, t)))

    xlsx = out_dir / "resultats.xlsx"
    assert xlsx.exists()

    wb = load_workbook(xlsx)
    assert "Synthèse" in wb.sheetnames
    assert "Comptes détectés" in wb.sheetnames
    assert "Newsletters" in wb.sheetnames
    assert "Tous les expéditeurs" in wb.sheetnames

    accounts_rows = list(wb["Comptes détectés"].iter_rows(values_only=True))
    accounts_headers = accounts_rows[0]
    accounts_data = accounts_rows[1:]
    accounts_by_domain = {row[0]: dict(zip(accounts_headers, row)) for row in accounts_data}

    newsletters_rows = list(wb["Newsletters"].iter_rows(values_only=True))
    newsletters_data = newsletters_rows[1:]
    newsletter_domains = {row[0] for row in newsletters_data}

    # --- Assertions adversariales ---

    # Amazon : compte (achats), pas newsletter
    assert "amazon.fr" in accounts_by_domain, "Amazon doit apparaître dans les comptes"
    assert accounts_by_domain["amazon.fr"]["Type"] == "🛒 Achats / facturation"
    assert "amazon.fr" not in newsletter_domains, "Amazon NE doit PAS être classé newsletter"

    # Substack : newsletter
    assert "substack.com" in newsletter_domains, "Substack doit être newsletter"

    # Mailchimp customer (Precedence bulk dès le 1er) : newsletter
    assert "mailchimp-customer.fr" in newsletter_domains

    # Snapple : ne doit PAS être marqué service connu (faux positif corrigé)
    snapple_in_accounts = "snapple.com" in accounts_by_domain
    if snapple_in_accounts:
        # Il est dans les comptes UNIQUEMENT s'il est newsletter-non + automated + count>=3
        # ici count=1, donc ne devrait pas être compte par cette voie.
        # En revanche, ⭐ Service connu doit être interdit.
        assert accounts_by_domain["snapple.com"]["Type"] != "⭐ Service connu"

    # Conversation humaine perso.fr : ne doit PAS être compte
    assert "perso.fr" not in accounts_by_domain, \
        "Conversation humaine ne doit pas être classée compte"

    # 2FA Google : compte (auth)
    assert "google.com" in accounts_by_domain
    assert accounts_by_domain["google.com"]["Type"] == "🔐 Authentification"

    # LinkedIn : compte service connu (subdomain digest normalisé en linkedin.com)
    assert "linkedin.com" in accounts_by_domain
    # Type peut être ⭐ Service connu OU 📬 Échanges répétés selon les autres signaux
    assert accounts_by_domain["linkedin.com"]["Type"] in (
        "⭐ Service connu", "🔐 Authentification", "📬 Échanges répétés"
    )

    # Notion : compte signup
    assert "notion.so" in accounts_by_domain
    assert accounts_by_domain["notion.so"]["Type"] == "🔑 Inscription trouvée"

    # Vérifie que la nouvelle colonne "Dernière activité" existe
    assert "Dernière activité" in accounts_headers
    assert "Exemples (sujets)" in accounts_headers
