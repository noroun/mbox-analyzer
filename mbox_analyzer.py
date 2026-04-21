"""
Analyse un export Gmail (Google Takeout) pour identifier
les comptes en ligne, newsletters et abonnements liés à une adresse email.

Génère un fichier Excel (resultats.xlsx) avec 3 onglets :
  - Comptes détectés     : sites où vous avez probablement créé un compte
  - Newsletters          : newsletters/abonnements avec lien de désinscription cliquable
  - Tous les expéditeurs : liste complète pour analyse manuelle
"""

import mailbox
import os
import re
import subprocess
import sys
import threading
import webbrowser
from collections import defaultdict
from datetime import datetime
from email.utils import parseaddr, parsedate_to_datetime, getaddresses
from email.header import decode_header, make_header
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# --------- Patterns de détection ---------

# --- Trois familles de signaux dans le sujet ---
# Inscription / activation / bienvenue → preuve forte qu'un compte a été créé.
SIGNUP_KEYWORDS = [
    "bienvenue", "welcome", "confirm", "vérifier votre", "verify your",
    "votre inscription", "your account", "votre compte",
    "activation", "activez votre", "activate your",
    "validation", "valider votre", "merci de votre inscription",
    "thanks for signing up", "thank you for registering",
    "créez votre compte", "create your account",
]

# Authentification : reset mdp, 2FA, alertes de connexion → compte actif.
AUTH_KEYWORDS = [
    "mot de passe", "password reset", "réinitialisation",
    "nouvelle connexion", "new sign-in", "new login",
    "code de vérification", "verification code", "code à 6 chiffres",
    "your login code", "sécurité de votre compte", "security alert",
    "alerte de connexion", "inactivité de votre compte",
    "connectez-vous", "sign in to",
]

# Transactionnel : commandes, factures, paiements → compte présent mais profil "achat".
TRANSACTION_KEYWORDS = [
    "votre commande", "your order", "receipt", "reçu",
    "facture", "invoice", "payment confirmation",
    "votre candidature", "your application", "candidature via",
]

# Mots-clés indiquant une newsletter / contenu marketing récurrent
NEWSLETTER_KEYWORDS = [
    "newsletter", "se désabonner", "unsubscribe", "désinscription",
    "vous recevez cet email", "you received this email",
    "ne plus recevoir", "manage your preferences", "gérer vos préférences",
    "modifier vos préférences", "update your preferences",
    "list-unsubscribe", "opt-out", "opt out",
]

# Domaines à ignorer (messageries perso)
IGNORED_DOMAINS = {
    "gmail.com", "googlemail.com", "yahoo.fr", "yahoo.com",
    "hotmail.fr", "hotmail.com", "outlook.fr", "outlook.com",
    "live.fr", "live.com", "wanadoo.fr", "free.fr", "orange.fr",
    "sfr.fr", "laposte.net", "aol.com", "icloud.com", "me.com",
    "protonmail.com", "proton.me", "tutanota.com",
}

# Labels Gmail à exclure (envoyés/brouillons/corbeille). Le spam est gardé
# mais marqué via is_spam — il sert de signal, pas de cause d'exclusion.
EXCLUDED_LABEL_KEYWORDS = [
    "messages envoyés", "sent", "brouillons", "drafts",
    "corbeille", "trash",
]
SPAM_LABEL_KEYWORDS = ["spam"]

# Services connus, comparés en match exact contre le domaine déjà normalisé
# (évite "apple" qui matchait "snapple.com" ou "google" qui matchait "googleads-fake.com")
SERVICE_DOMAINS = {
    "indeed.com", "indeed.fr", "linkedin.com", "twitter.com", "x.com",
    "facebook.com", "instagram.com", "tiktok.com", "youtube.com",
    "amazon.com", "amazon.fr", "ebay.com", "ebay.fr", "paypal.com",
    "stripe.com", "booking.com", "airbnb.com", "airbnb.fr",
    "uber.com", "ubereats.com", "deliveroo.fr", "deliveroo.com",
    "doctolib.fr", "doctolib.com", "github.com", "gitlab.com",
    "bitbucket.org", "slack.com", "notion.so", "figma.com",
    "dropbox.com", "google.com", "microsoft.com", "apple.com",
    "adobe.com", "spotify.com", "netflix.com", "deezer.com",
    "twitch.tv", "discord.com", "shopify.com", "wordpress.com",
    "medium.com", "substack.com",
}


# Mapping de regroupement : sous-domaine → domaine principal
# (évite de compter accounts.google.com et google.com séparément)
def normalize_domain(domain):
    """Regroupe les sous-domaines connus sous leur domaine principal."""
    if not domain:
        return domain
    parts = domain.split(".")
    # Garde les 2 derniers segments sauf cas particuliers (co.uk, etc.)
    if len(parts) >= 3:
        # Cas .co.uk, .com.au, .co.jp, etc.
        if parts[-2] in ("co", "com", "net", "org", "gov") and len(parts[-1]) == 2:
            return ".".join(parts[-3:])
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return domain


# Regex pour extraire les liens de désinscription dans le corps d'un mail.
# Le header List-Unsubscribe est parsé séparément (cf. find_unsubscribe_link).
UNSUB_BODY_PATTERN = re.compile(
    r'https?://[^\s<>"\'\)]+(?:unsubscribe|desinscription|desabonn|opt[-_]?out|preferences|mailing-?list)[^\s<>"\'\)]*',
    re.I,
)


def decode_mime_header(value):
    """Décode un header MIME-encodé (ex: =?utf-8?q?V=C3=A9rifier?=) en string lisible."""
    if not value:
        return ""
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return value


def _decode_payload(payload, charset):
    """Décode un payload bytes en str avec fallback robuste sur les charsets foireux."""
    if not payload:
        return ""
    candidates = []
    if charset:
        candidates.append(charset)
    candidates.extend(["utf-8", "latin-1"])
    for cs in candidates:
        try:
            return payload.decode(cs, errors="strict")
        except (LookupError, UnicodeDecodeError):
            continue
    # Dernier recours : on remplace les octets invalides plutôt que de tout perdre.
    return payload.decode("utf-8", errors="replace")


def get_email_body(message):
    """Extrait le corps texte d'un mail (gère multipart)."""
    body = ""
    if message.is_multipart():
        for part in message.walk():
            ctype = part.get_content_type()
            if ctype in ("text/plain", "text/html"):
                try:
                    payload = part.get_payload(decode=True)
                    body += _decode_payload(payload, part.get_content_charset())
                except Exception:
                    pass
    else:
        try:
            payload = message.get_payload(decode=True)
            body = _decode_payload(payload, message.get_content_charset())
        except Exception:
            pass
    return body


# Locale-part patterns indicating an automated/transactional sender
# (utilisé pour distinguer un échange humain d'un service)
_AUTOMATED_LOCAL_RE = re.compile(
    r"^(?:no[-_]?reply|do[-_]?not[-_]?reply|donotreply|noreply|"
    r"postmaster|notifications?|alerts?|news|info|hello|"
    r"support|service|automated|mailer|contact|bonjour|"
    r"confirmation|noresponse|reply)(?:[-_.+]|$)",
    re.I,
)


def _open_file_with_default_app(path):
    """Ouvre un fichier avec l'application par défaut de l'OS. Renvoie True si le
    lancement a réussi (ne dit rien sur le succès de l'ouverture par l'app cible)."""
    path = str(path)
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", path], check=True)
        elif sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.run(["xdg-open", path], check=True)
        return True
    except Exception:
        return False


def is_automated_sender(email_addr):
    """True si l'adresse ressemble à un expéditeur automatisé (noreply, notifications, ...)."""
    if not email_addr or "@" not in email_addr:
        return False
    local = email_addr.split("@", 1)[0]
    return bool(_AUTOMATED_LOCAL_RE.match(local))


def find_unsubscribe_link(message, body):
    """Cherche un lien de désinscription dans le header List-Unsubscribe ou le corps."""
    # 1. Header List-Unsubscribe (standard)
    list_unsub = message.get("List-Unsubscribe", "")
    if list_unsub:
        urls = re.findall(r'<(https?://[^>]+)>', list_unsub)
        if urls:
            return urls[0]
    # 2. Recherche dans le corps
    match = UNSUB_BODY_PATTERN.search(body)
    if match:
        return match.group(0)
    return ""


def analyse(mbox_path, output_dir, progress_callback):
    """Parse le MBOX et génère les 3 fichiers CSV."""

    senders = defaultdict(lambda: {
        "count": 0,
        "first": None,
        "last": None,
        "signup_subjects": set(),
        "auth_subjects": set(),
        "transaction_subjects": set(),
        "is_newsletter": False,
        "unsub_link": "",
        "_pending_unsub_link": "",  # candidat header List-Unsubscribe en attente de corroboration
        "sender_name": "",
        "sender_email": "",
        "is_service": False,
        "is_automated": False,
        "in_spam": False,
    })

    mbox = mailbox.mbox(mbox_path)
    total = len(mbox)
    progress_callback(0, total)

    # --- Détection auto des adresses de l'utilisateur ---
    # On échantillonne jusqu'à 5000 mails et on collecte TOUS les destinataires
    # (Delivered-To, To, Cc, Bcc, X-Original-To). On normalise les +tag en base.
    # On ne garde que les adresses récurrentes : >=5 % de l'échantillon ou >=10 occurrences.
    SAMPLE_CAP = 5000
    sample_n = min(SAMPLE_CAP, total)
    addr_counts = defaultdict(int)
    recipient_headers = ("Delivered-To", "To", "Cc", "Bcc", "X-Original-To")
    for idx, msg in enumerate(mbox):
        if idx >= sample_n:
            break
        pairs = []
        for h in recipient_headers:
            v = msg.get_all(h)
            if v:
                pairs.extend(getaddresses(v))
        for _, addr in pairs:
            if not addr or "@" not in addr:
                continue
            addr = addr.lower().strip()
            local, _, dom = addr.partition("@")
            if "+" in local:
                local = local.split("+", 1)[0]
            addr = f"{local}@{dom}"
            addr_counts[addr] += 1

    threshold = max(10, sample_n // 20)  # >=5 % de l'échantillon ou >=10
    user_emails = {a for a, c in addr_counts.items() if c >= threshold}

    skipped_sent = 0
    skipped_spam = 0
    total_after_filter = 0

    for i, message in enumerate(mbox):
        if i % 200 == 0:
            progress_callback(i, total)

        # Lire les labels Gmail (MIME-décodé)
        labels_raw = message.get("X-Gmail-Labels", "") or ""
        labels = decode_mime_header(labels_raw).lower()

        # Exclusion : mails envoyés, brouillons, corbeille
        if any(kw in labels for kw in EXCLUDED_LABEL_KEYWORDS):
            skipped_sent += 1
            continue

        # Flag spam (on le garde mais on le marque)
        is_spam = any(kw in labels for kw in SPAM_LABEL_KEYWORDS)

        from_header = message.get("From", "") or ""
        subject_raw = message.get("Subject", "") or ""
        subject = decode_mime_header(subject_raw).strip()
        date_header = message.get("Date", "") or ""

        name_raw, email_addr = parseaddr(from_header)
        name = decode_mime_header(name_raw)
        if not email_addr or "@" not in email_addr:
            continue

        email_addr_lower = email_addr.lower()

        # Exclusion : mails auto-envoyés (expéditeur = destinataire). Normalise +tag.
        _local, _, _dom = email_addr_lower.partition("@")
        if "+" in _local:
            _local = _local.split("+", 1)[0]
        if f"{_local}@{_dom}" in user_emails:
            skipped_sent += 1
            continue

        raw_domain = email_addr_lower.split("@")[-1].strip()
        if not raw_domain or raw_domain in IGNORED_DOMAINS:
            continue

        # Normalisation : accounts.google.com → google.com
        domain = normalize_domain(raw_domain)

        total_after_filter += 1

        try:
            date = parsedate_to_datetime(date_header) if date_header else None
        except Exception:
            date = None

        s = senders[domain]
        s["count"] += 1
        if not s["sender_email"]:
            s["sender_email"] = email_addr
        if name and not s["sender_name"]:
            s["sender_name"] = name
        if is_spam:
            s["in_spam"] = True

        if date:
            if not s["first"] or date.replace(tzinfo=None) < s["first"].replace(tzinfo=None):
                s["first"] = date
            if not s["last"] or date.replace(tzinfo=None) > s["last"].replace(tzinfo=None):
                s["last"] = date

        # Détection service connu (match exact sur le domaine déjà normalisé)
        if domain in SERVICE_DOMAINS:
            s["is_service"] = True

        # Expéditeur automatisé ?
        if is_automated_sender(email_addr_lower):
            s["is_automated"] = True

        subject_lower = subject.lower()

        # Détection des 3 familles de signaux dans le sujet
        # Note: substring match volontaire — "confirm" doit matcher "confirmation"
        # (forme française courante). Frontières de mots écarteraient ce cas.
        has_signup_kw = any(kw in subject_lower for kw in SIGNUP_KEYWORDS)
        has_auth_kw = any(kw in subject_lower for kw in AUTH_KEYWORDS)
        has_transaction_kw = any(kw in subject_lower for kw in TRANSACTION_KEYWORDS)

        if has_signup_kw and len(s["signup_subjects"]) < 3:
            s["signup_subjects"].add(subject[:100])
        if has_auth_kw and len(s["auth_subjects"]) < 3:
            s["auth_subjects"].add(subject[:100])
        if has_transaction_kw and len(s["transaction_subjects"]) < 3:
            s["transaction_subjects"].add(subject[:100])

        # --- Détection newsletter avec corroboration ---
        # Un mail avec List-Unsubscribe ne suffit pas (Amazon le met sur les
        # confirmations de commande). On exige au moins un signal additionnel.
        if not s["is_newsletter"]:
            if message.get("List-Unsubscribe"):
                precedence = (message.get("Precedence", "") or "").strip().lower()
                auto_submitted = (message.get("Auto-Submitted", "") or "").strip().lower()
                bulk_signal = precedence.startswith(("bulk", "list"))
                auto_signal = bool(auto_submitted) and auto_submitted != "no"
                repeat_signal = s["count"] >= 2
                non_transactional = not has_transaction_kw

                if bulk_signal or auto_signal or (repeat_signal and non_transactional):
                    s["is_newsletter"] = True
                    s["unsub_link"] = find_unsubscribe_link(message, "")
                else:
                    # Mémorise le lien au cas où un futur mail du même domaine corrobore.
                    if not s["_pending_unsub_link"]:
                        s["_pending_unsub_link"] = find_unsubscribe_link(message, "")
            elif s["count"] <= 3:  # n'analyse le corps que pour les 3 premiers mails par domaine (perf)
                body = get_email_body(message)
                if any(kw in body.lower() for kw in NEWSLETTER_KEYWORDS):
                    s["is_newsletter"] = True
                    s["unsub_link"] = find_unsubscribe_link(message, body)

    progress_callback(total, total)

    # --------- Export XLSX ---------
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "resultats.xlsx"

    def fmt_date(d):
        return d.strftime("%Y-%m-%d") if d else ""

    _now_naive = datetime.now()

    def months_since(d):
        """Renvoie une étiquette d'ancienneté lisible : '3 mois', '2 ans', '—' si inconnue."""
        if not d:
            return "—"
        try:
            d_naive = d.replace(tzinfo=None)
        except Exception:
            return "—"
        delta_days = (_now_naive - d_naive).days
        if delta_days < 0:
            return "—"
        months = delta_days // 30
        if months < 1:
            return "< 1 mois"
        if months < 12:
            return f"{months} mois"
        years = months // 12
        rem_months = months % 12
        if rem_months == 0:
            return f"{years} an{'s' if years > 1 else ''}"
        return f"{years} an{'s' if years > 1 else ''} {rem_months} mois"

    def months_since_int(d):
        if not d:
            return None
        try:
            return (_now_naive - d.replace(tzinfo=None)).days // 30
        except Exception:
            return None

    # Logique : un domaine est un "compte probable" si...
    def is_probable_account(s):
        return (
            bool(s["signup_subjects"])              # inscription trouvée
            or bool(s["auth_subjects"])             # reset mdp / 2FA / alerte connexion
            or bool(s["transaction_subjects"])      # commande / facture
            or s["is_service"]                      # service connu (LinkedIn, Indeed, etc.)
            # Échanges répétés : on durcit en exigeant un expéditeur automatisé
            # pour éviter de classer les conversations humaines comme comptes.
            or (s["count"] >= 3 and not s["is_newsletter"] and s["is_automated"])
        )

    def account_type(s):
        """Étiquette dominante d'un compte selon les preuves les plus fortes."""
        if s["signup_subjects"]:
            return "🔑 Inscription trouvée"
        if s["auth_subjects"]:
            return "🔐 Authentification"
        if s["transaction_subjects"]:
            return "🛒 Achats / facturation"
        if s["is_service"]:
            return "⭐ Service connu"
        return "📬 Échanges répétés"

    def example_subjects(s):
        """Renvoie jusqu'à 3 sujets représentatifs du bucket le plus pertinent."""
        for bucket in ("signup_subjects", "auth_subjects", "transaction_subjects"):
            if s[bucket]:
                return " | ".join(s[bucket])
        return ""

    # Styles
    FONT = "Century Gothic"
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E7D32")
    cell_font = Font(name=FONT, size=10)
    link_font = Font(name=FONT, size=10, color="1565C0", underline="single")
    spam_font = Font(name=FONT, size=10, color="C62828", italic=True)
    alt_fill = PatternFill("solid", start_color="F5F5F5")
    spam_fill = PatternFill("solid", start_color="FFEBEE")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def style_data_rows(ws, rows_data, ncols, numeric_cols=(), link_cols=()):
        """rows_data = liste de tuples (row_number, is_spam) pour styler selon."""
        for row_num, is_spam in rows_data:
            fill = spam_fill if is_spam else (alt_fill if row_num % 2 == 0 else None)
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row_num, column=col)
                if col in link_cols and cell.value:
                    cell.font = link_font
                elif is_spam:
                    cell.font = spam_font
                else:
                    cell.font = cell_font
                cell.alignment = center if col in numeric_cols else left
                cell.border = thin_border
                if fill:
                    cell.fill = fill

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Préparation des listes
    accounts = sorted(
        [(d, s) for d, s in senders.items() if is_probable_account(s)],
        key=lambda x: -x[1]["count"],
    )
    newsletters = sorted(
        [(d, s) for d, s in senders.items() if s["is_newsletter"]],
        key=lambda x: -x[1]["count"],
    )
    all_senders = sorted(senders.items(), key=lambda x: -x[1]["count"])
    spam_count = sum(1 for _, s in senders.items() if s["in_spam"])

    wb = Workbook()

    # ========== Onglet 1 : Synthèse ==========
    ws0 = wb.active
    ws0.title = "Synthèse"
    synth_font_title = Font(name=FONT, bold=True, size=14, color="1B5E20")
    synth_font_h2 = Font(name=FONT, bold=True, size=12, color="2E7D32")
    synth_font_body = Font(name=FONT, size=11)
    synth_font_big = Font(name=FONT, bold=True, size=18, color="2E7D32")

    ws0["A1"] = "📧 Analyse de votre boîte Gmail"
    ws0["A1"].font = synth_font_title
    ws0.merge_cells("A1:D1")

    ws0["A3"] = "Chiffres clés"
    ws0["A3"].font = synth_font_h2

    inactive_24m = sum(
        1 for _, s in accounts
        if (m := months_since_int(s["last"])) is not None and m > 24
    )
    top10_count = min(10, len(accounts))

    stats = [
        ("Messages analysés", total_after_filter),
        ("Messages envoyés / brouillons ignorés", skipped_sent),
        ("Expéditeurs uniques", len(senders)),
        ("Comptes probables à traiter", len(accounts)),
        ("Newsletters à résilier", len(newsletters)),
        ("Comptes inactifs (> 24 mois)", inactive_24m),
        ("Comptes très actifs (top 10 par volume)", top10_count),
        ("Domaines présents dans le Spam", spam_count),
    ]
    for i, (label, val) in enumerate(stats, start=4):
        ws0.cell(row=i, column=1, value=label).font = synth_font_body
        c = ws0.cell(row=i, column=2, value=val)
        c.font = synth_font_big
        c.alignment = Alignment(horizontal="right")

    ws0.cell(row=4 + len(stats) + 1, column=1, value="Prochaines étapes").font = synth_font_h2
    steps = [
        "1. Ouvrez l'onglet « Comptes détectés » → changez l'email de connexion sur chaque site",
        "2. Ouvrez l'onglet « Newsletters » → cliquez sur les liens de désinscription pour chaque abonnement inutile",
        "3. Consultez « Tous les expéditeurs » si un site manque — l'outil ne détecte pas tous les comptes automatiquement",
        "4. Priorisez les services critiques : banque, impôts, assurance, Apple ID / Google, réseaux sociaux",
        "5. Avant de fermer la boîte, mettez en place une redirection 6-12 mois vers votre nouvelle adresse",
    ]
    for i, step in enumerate(steps, start=4 + len(stats) + 2):
        c = ws0.cell(row=i, column=1, value=step)
        c.font = synth_font_body
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        ws0.row_dimensions[i].height = 20

    ws0.column_dimensions["A"].width = 50
    ws0.column_dimensions["B"].width = 15
    ws0.column_dimensions["C"].width = 15
    ws0.column_dimensions["D"].width = 15

    # ========== Onglet 2 : Comptes détectés ==========
    ws1 = wb.create_sheet("Comptes détectés")
    headers1 = ["Domaine", "Expéditeur", "Nb mails", "Premier mail", "Dernier mail",
                "Dernière activité", "Type", "Exemples (sujets)"]
    ws1.append(headers1)
    rows_data_1 = []
    for d, s in accounts:
        ws1.append([
            d,
            s["sender_name"] or s["sender_email"],
            s["count"],
            fmt_date(s["first"]),
            fmt_date(s["last"]),
            months_since(s["last"]),
            account_type(s),
            example_subjects(s),
        ])
        rows_data_1.append((ws1.max_row, s["in_spam"]))
    style_header(ws1, len(headers1))
    style_data_rows(ws1, rows_data_1, len(headers1), numeric_cols=(3, 4, 5, 6, 7))
    autosize(ws1, [25, 30, 12, 14, 14, 18, 22, 55])
    if accounts:
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(accounts) + 1}"

    # ========== Onglet 3 : Newsletters ==========
    ws2 = wb.create_sheet("Newsletters")
    headers2 = ["Domaine", "Expéditeur", "Nb mails", "Dernier mail", "Lien de désinscription"]
    ws2.append(headers2)
    rows_data_2 = []
    for d, s in newsletters:
        ws2.append([d, s["sender_name"] or s["sender_email"], s["count"], fmt_date(s["last"]), s["unsub_link"]])
        if s["unsub_link"]:
            cell = ws2.cell(row=ws2.max_row, column=5)
            cell.hyperlink = s["unsub_link"]
        rows_data_2.append((ws2.max_row, s["in_spam"]))
    style_header(ws2, len(headers2))
    style_data_rows(ws2, rows_data_2, len(headers2), numeric_cols=(3, 4), link_cols=(5,))
    autosize(ws2, [25, 30, 12, 14, 70])
    if newsletters:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(newsletters) + 1}"

    # ========== Onglet 4 : Tous les expéditeurs ==========
    ws3 = wb.create_sheet("Tous les expéditeurs")
    headers3 = ["Domaine", "Expéditeur", "Nb mails", "Premier mail", "Dernier mail",
                "Compte probable", "Newsletter", "Dans Spam"]
    ws3.append(headers3)
    rows_data_3 = []
    for d, s in all_senders:
        ws3.append([
            d,
            s["sender_name"] or s["sender_email"],
            s["count"],
            fmt_date(s["first"]),
            fmt_date(s["last"]),
            "✅" if is_probable_account(s) else "",
            "✅" if s["is_newsletter"] else "",
            "⚠️" if s["in_spam"] else "",
        ])
        rows_data_3.append((ws3.max_row, s["in_spam"]))
    style_header(ws3, len(headers3))
    style_data_rows(ws3, rows_data_3, len(headers3), numeric_cols=(3, 4, 5, 6, 7, 8))
    autosize(ws3, [25, 30, 12, 14, 14, 16, 12, 12])
    if all_senders:
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{len(all_senders) + 1}"

    # ========== Onglet Aide (en premier) ==========
    ws_instr = wb.create_sheet("Aide", 0)

    instr_title_font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    instr_title_fill = PatternFill("solid", start_color="1B5E20")
    instr_h2_font = Font(name=FONT, bold=True, size=12, color="1B5E20")
    instr_body_font = Font(name=FONT, size=11)
    instr_note_font = Font(name=FONT, size=10, italic=True, color="555555")
    instr_prio_fill = PatternFill("solid", start_color="E8F5E9")
    instr_warn_fill = PatternFill("solid", start_color="FFF9C4")

    def _instr_row(r, text, font=None, fill=None, indent=0, height=18):
        cell = ws_instr.cell(row=r, column=1)
        cell.value = ("    " * indent) + text
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_instr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws_instr.row_dimensions[r].height = height
        return r + 1

    instructions = [
        ("📋  INSTRUCTIONS — Actions à mener avant de fermer votre Gmail",
         instr_title_font, instr_title_fill, 0, 32),
        ("", instr_body_font, None, 0, 8),

        ("1.  COMPTES DÉTECTÉS  →  onglet « Comptes détectés »",
         instr_h2_font, instr_prio_fill, 0, 22),
        ("Connectez-vous sur chaque site et mettez à jour l'adresse email de connexion.",
         instr_body_font, None, 1, 18),
        ("Ordre de priorité : administratif (impôts, CAF…) > banque/finance > pro (LinkedIn, GitHub…) > réseaux sociaux > loisirs > reste.",
         instr_body_font, None, 1, 20),
        ("Comptes inactifs (> 2 ans) : envisagez la suppression du compte plutôt que le simple changement d'email.",
         instr_body_font, None, 1, 20),
        ("Certains comptes peuvent être abandonnés (services morts ou sans intérêt) — c'est tout à fait acceptable.",
         instr_body_font, None, 1, 18),
        ("", instr_body_font, None, 0, 6),

        ("2.  NEWSLETTERS  →  onglet « Newsletters »",
         instr_h2_font, instr_prio_fill, 0, 22),
        ("Cliquez sur le lien dans la colonne « Lien de désinscription » pour vous désabonner en un clic.",
         instr_body_font, None, 1, 20),
        ("Alternative : connectez-vous au site et modifiez vos préférences e-mail.",
         instr_body_font, None, 1, 18),
        ("", instr_body_font, None, 0, 6),

        ("3.  COMPLÉTEZ L'ANALYSE — 2 pages Google indispensables",
         instr_h2_font, instr_prio_fill, 0, 22),
        ("Ces comptes sont invisibles dans le MBOX (comptes SSO ou mots de passe enregistrés).",
         instr_note_font, None, 1, 18),
        ("passwords.google.com  →  tous les sites avec un mot de passe Chrome/Android.",
         instr_body_font, None, 1, 18),
        ("myaccount.google.com/connections  →  services connectés via « Se connecter avec Google ».",
         instr_body_font, None, 1, 18),
        ("", instr_body_font, None, 0, 6),

        ("4.  DOMAINES EN SPAM  (fond rouge dans les onglets)",
         instr_h2_font, instr_warn_fill, 0, 22),
        ("En général : inutile de contacter ces services (spam/phishing probable).",
         instr_body_font, None, 1, 18),
        ("Vérifiez quand même si vous avez un compte légitime chez eux.",
         instr_body_font, None, 1, 18),
        ("", instr_body_font, None, 0, 6),

        ("5.  AVANT DE FERMER VOTRE BOÎTE GMAIL",
         instr_h2_font, instr_prio_fill, 0, 22),
        ("Faites-le EN PREMIER : activez la redirection vers votre nouvelle adresse (Paramètres Gmail → Transfert et POP/IMAP). Vous continuerez à recevoir vos mails pendant toute la migration, même si vous oubliez un compte.",
         instr_body_font, None, 1, 32),
        ("Attendez d'avoir traité TOUS les comptes critiques avant de supprimer la boîte.",
         instr_body_font, None, 1, 18),
        ("Prévenez vos contacts importants de votre nouvelle adresse.",
         instr_body_font, None, 1, 18),
    ]

    row_idx = 1
    for text, font, fill, indent, height in instructions:
        row_idx = _instr_row(row_idx, text, font=font, fill=fill, indent=indent, height=height)

    # Liens cliquables pour les deux pages Google
    link_font_instr = Font(name=FONT, size=11, color="1565C0", underline="single")
    for r in range(1, row_idx):
        cell = ws_instr.cell(row=r, column=1)
        val = str(cell.value or "")
        if "passwords.google.com" in val:
            cell.hyperlink = "https://passwords.google.com"
            cell.font = link_font_instr
        elif "myaccount.google.com/connections" in val:
            cell.hyperlink = "https://myaccount.google.com/connections"
            cell.font = link_font_instr

    ws_instr.column_dimensions["A"].width = 85
    ws_instr.column_dimensions["B"].width = 5
    ws_instr.column_dimensions["C"].width = 5
    ws_instr.column_dimensions["D"].width = 5

    wb.save(xlsx_path)

    return {
        "total_domaines": len(senders),
        "comptes": len(accounts),
        "newsletters": len(newsletters),
        "analyses": total_after_filter,
        "ignores": skipped_sent,
        "spam": spam_count,
    }


# --------- Contenu de l'aide (rendu dans une fenêtre Tk) ---------

# HELP_CONTENT est une transcription fidèle de README.md. Si tu modifies le README,
# mets à jour cette liste en conséquence (voir test_help_matches_readme).
HELP_CONTENT = [
    ("title", "📧 Zoquez Google"),
    ("subtitle", "Identifiez tous les sites où vous avez un compte lié à votre adresse gadz.org."),
    ("p", "⚠️ Cet outil est uniquement à titre informatif. Le plus important est d'avoir activé la redirection et de suivre les instructions de l'équipe gadz.org et de vos DDP. Consultez ces instructions en priorité — cet outil vient en complément, pas en remplacement."),
   ("p", " Activez la redirection en premier — Paramètres Gmail → Transfert et POP/IMAP. Vous continuerez à recevoir vos mails pendant toute la migration, même si vous oubliez un compte."),

    ("p", "Cet outil analyse l'export de votre boîte mail Gmail et génère un fichier Excel (resultats.xlsx) avec trois onglets :"),
    ("ul", [
        "✅ Comptes détectés — sites où vous avez probablement créé un compte",
        "📰 Newsletters — abonnements avec lien de désinscription cliquable",
        "📋 Tous les expéditeurs — liste complète pour vérification manuelle",
    ]),
    ("sep", None),

    ("h2", "🔒 Confidentialité"),
    ("ul", [
        "✅ Tout est traité localement sur votre ordinateur. Aucune donnée n'est envoyée sur Internet.",
        "✅ Le code est ouvert, vous pouvez le vérifier dans mbox_analyzer.py.",
        "✅ L'outil ne lit que votre fichier MBOX, ne se connecte à rien.",
    ]),

    ("sep", None),

    ("h2", "🚀 Tutoriel pas à pas (aucune connaissance technique requise)"),

    ("h3", "Étape 1 — Exportez votre Gmail"),
    ("ol", [
        "Allez sur https://takeout.google.com",
        "Cliquez sur « Tout désélectionner » en haut",
        "Faites défiler la liste et cochez uniquement « Messagerie » (Mail en anglais)",
        "Tout en bas, cliquez sur « Étape suivante »",
        "Choisissez : Méthode de livraison = Envoyer le lien de téléchargement par e-mail ; Type d'export = Une seule exportation ; Type de fichier = .zip ; Taille maximale = 2 Go (suffisant dans la plupart des cas)",
        "Cliquez sur « Créer l'exportation »",
    ]),
    ("note", "⏱️ Google met de quelques heures à plusieurs jours à préparer votre archive. Vous recevrez un email avec le lien de téléchargement quand c'est prêt."),

    ("h3", "Étape 2 — Récupérez le fichier MBOX"),
    ("ol", [
        "Téléchargez le ZIP que Google vous a envoyé",
        "Décompressez-le",
        "Vous trouverez un fichier .mbox dans le dossier Takeout/Messagerie/ — c'est lui qu'il faut analyser",
    ]),

    ("h3", "Étape 3 — Lancez l'analyse"),
    ("h3", "Sur Mac"),
    ("ol", [
        "Double-cliquez sur ZoquerGMail.app",
        "⚠️ Si Mac affiche « Apple ne peut pas vérifier que cette app ne contient pas de logiciel malveillant » : faites clic droit sur l'app → Ouvrir → Ouvrir dans la fenêtre de confirmation. Vous n'aurez à le faire qu'une seule fois.",
    ]),
    ("h3", "Sur Windows"),
    ("ol", [
        "Double-cliquez sur ZoquerGMail.exe",
        "⚠️ Si Windows affiche « Windows a protégé votre ordinateur » : cliquez sur « Informations complémentaires », puis sur « Exécuter quand même ».",
    ]),

    ("h3", "Étape 4 — Utilisez l'outil"),
    ("ol", [
        "Cliquez sur « Parcourir » à côté de « Fichier MBOX » et sélectionnez votre fichier .mbox",
        "Le dossier de sortie est rempli automatiquement (vous pouvez le changer)",
        "Cliquez sur « Lancer l'analyse »",
        "⏱️ Patientez : selon la taille de votre boîte, ça peut prendre de quelques minutes à 30 minutes",
    ]),

    ("h3", "Étape 5 — Exploitez les résultats"),
    ("p", "L'outil génère un fichier resultats.xlsx ouvrable dans Excel, Numbers, ou Google Sheets."),
  
    ("ul", [
        "Priorisez par ordre d'importance : administratif (impôts, CAF…) → banque/finance → pro (LinkedIn, GitHub…) → réseaux sociaux → loisirs → reste.",
        "Acceptez de perdre certains comptes : les services morts ou sans intérêt n'ont pas besoin d'être traités — c'est tout à fait normal.",
    ]),

    ("sep", None),

    ("h2", "🔍 Compléter avec deux pages Google indispensables"),
    ("p", "L'analyse du MBOX rate par construction deux types de comptes : ceux dont vous avez supprimé les mails de bienvenue, et ceux ouverts via « Se connecter avec Google » (SSO/OAuth) qui ne laissent souvent aucune trace dans la boîte. Pour les retrouver, ouvrez ces deux pages dans votre navigateur (connecté à votre compte Google) :"),
    ("ul", [
        "🔑 Mots de passe enregistrés → https://passwords.google.com — la liste de tous les sites pour lesquels Chrome / Android a stocké un mot de passe. Recoupez avec l'onglet Comptes détectés — tout site présent ici et absent du XLSX est un compte à traiter.",
        "🔗 Applications connectées via Google → https://myaccount.google.com/connections — la liste des services tiers connectés via Sign in with Google (Spotify, Notion, Canva, Figma, etc.). C'est la seule façon fiable de retrouver les comptes SSO, invisibles dans les emails.",
    ]),
    ("p", "Pour chaque entrée trouvée sur ces deux pages : changez l'email/identifiant de connexion vers votre nouvelle adresse, puis révoquez l'accès Google si vous fermez votre compte Gmail."),

    ("sep", None),

    ("h2", "❓ FAQ"),
    ("h3", "Q : Combien de temps prend l'analyse ?"),
    ("p", "R : Environ 1 minute par 10 000 mails. Une boîte de 50 000 mails ≈ 5 minutes."),
    ("h3", "Q : Mes données sont-elles envoyées quelque part ?"),
    ("p", "R : Non, jamais. Tout reste sur votre ordinateur."),
    ("h3", "Q : Pourquoi certains comptes ne sont pas détectés ?"),
    ("p", "R : L'outil détecte les comptes via les mails d'inscription/bienvenue/réinitialisation. Si vous avez supprimé ces mails, le compte n'apparaîtra que dans tous_expediteurs.csv."),
    ("h3", "Q : Et pour Yahoo / Outlook / autre ?"),
    ("p", "R : Cet outil fonctionne avec n'importe quel fichier MBOX. Pour Outlook, exportez d'abord en MBOX via Thunderbird."),
    ("h3", "Q : À quoi ça sert ?"),
    ("p", "R : À queud's."),

    ("sep", None),

    ("h2", "🛠️ Pour les développeurs"),
    ("h3", "Lancer depuis les sources"),
    ("code", "git clone https://github.com/VOTRE_PSEUDO/mbox-analyzer\n"
             "cd mbox-analyzer\n"
             "pip install -r requirements.txt\n"
             "python3 mbox_analyzer.py"),
    ("p", "Seule dépendance externe : openpyxl (pour générer le fichier Excel formaté)."),

    ("h3", "Compiler localement"),
    ("code", "pip install -r requirements.txt\n"
             "pip install pyinstaller\n"
             "pyinstaller --onefile --windowed --name ZoquerGMail mbox_analyzer.py"),
    ("p", "L'exécutable apparaît dans dist/."),

    ("sep", None),

    ("h2", "📄 Licence"),
    ("p", "MIT — utilisez, modifiez et partagez librement."),
]


_URL_RE = re.compile(r"https?://[^\s)]+[^\s).,]", re.I)


def _render_text_with_links(parent, text, font, fg="#333", wraplength=580, pady=(0, 5), bg="white"):
    """Rend du texte où les URLs http(s) deviennent des liens cliquables (ouverture
    dans le navigateur). Si aucune URL, on garde un simple Label (plus léger)."""
    if not _URL_RE.search(text):
        tk.Label(parent, text=text, font=font, bg=bg, fg=fg,
                 anchor="w", justify="left",
                 wraplength=wraplength).pack(anchor="w", pady=pady)
        return

    # Largeur en caractères approximative (≈ wraplength / 7 px par caractère).
    char_width = max(40, wraplength // 7)
    # Comptage de lignes (approx) pour dimensionner le Text — on ajoute de la marge.
    approx_lines = max(1, len(text) // char_width + text.count("\n") + 1)
    widget = tk.Text(parent, wrap="word", borderwidth=0, highlightthickness=0,
                     font=font, bg=bg, fg=fg,
                     width=char_width, height=approx_lines,
                     cursor="arrow", takefocus=0)
    widget.pack(anchor="w", fill="x", pady=pady)

    # Insertion avec tags "link" sur chaque URL.
    pos = 0
    for match in _URL_RE.finditer(text):
        widget.insert("end", text[pos:match.start()])
        link_start = widget.index("end")   # position d'insertion AVANT le lien
        widget.insert("end", match.group(0))
        link_end = widget.index("end")     # position APRÈS le dernier char du lien
        tag = f"link-{match.start()}"
        widget.tag_add(tag, link_start, link_end)
        widget.tag_config(tag, foreground="#1565C0", underline=True)
        url = match.group(0)
        widget.tag_bind(tag, "<Button-1>", lambda _e, u=url: webbrowser.open(u))
        widget.tag_bind(tag, "<Enter>", lambda _e: widget.config(cursor="hand2"))
        widget.tag_bind(tag, "<Leave>", lambda _e: widget.config(cursor="arrow"))
        pos = match.end()
    widget.insert("end", text[pos:])

    # Ajuste la hauteur réelle en fonction des lignes rendues.
    widget.update_idletasks()
    real_lines = int(widget.index("end-1c").split(".")[0])
    widget.config(height=real_lines)
    # Lecture seule : on intercepte les éditions clavier mais pas les clics de tags.
    widget.bind("<Key>", lambda _e: "break")


def show_help(parent):
    """Affiche une fenêtre d'aide avec le contenu formaté."""
    win = tk.Toplevel(parent)
    win.title("Aide")
    win.geometry("820x700")
    win.transient(parent)

    BG = "#F0F0F0"

    # Conteneur avec scrollbar
    container = tk.Frame(win, bg=BG)
    container.pack(fill="both", expand=True)

    canvas = tk.Canvas(container, bg=BG, highlightthickness=0)
    scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    content = tk.Frame(canvas, bg=BG)

    content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=content, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Scroll molette (compatible macOS trackpad et souris)
    def on_mousewheel(event):
        delta = event.delta
        # Sur macOS le trackpad envoie de petites valeurs (pas multiples de 120)
        if abs(delta) >= 120:
            delta = int(-1 * (delta / 120))
        else:
            delta = int(-1 * delta) if delta != 0 else 0
        canvas.yview_scroll(delta, "units")
    canvas.bind_all("<MouseWheel>", on_mousewheel)
    canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
    canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))
    win.bind("<Destroy>", lambda e: (
        canvas.unbind_all("<MouseWheel>"),
        canvas.unbind_all("<Button-4>"),
        canvas.unbind_all("<Button-5>"),
    ))

    # Rendu du contenu
    inner = tk.Frame(content, bg=BG, padx=30, pady=20)
    inner.pack(fill="both", expand=True)

    for kind, value in HELP_CONTENT:
        if kind == "title":
            tk.Label(inner, text=value, font=(GOTHIC_FONT, 22, "bold"),
                     bg=BG, fg="#1B5E20", anchor="w", justify="left").pack(anchor="w", pady=(0, 5))
        elif kind == "subtitle":
            tk.Label(inner, text=value, font=(GOTHIC_FONT, 14, "italic"),
                     bg=BG, fg="#555", anchor="w", justify="left",
                     wraplength=600).pack(anchor="w", pady=(0, 15))
        elif kind == "h2":
            tk.Label(inner, text=value, font=(GOTHIC_FONT, 17, "bold"),
                     bg=BG, fg="#2E7D32", anchor="w", justify="left").pack(anchor="w", pady=(15, 8))
        elif kind == "h3":
            tk.Label(inner, text=value, font=(GOTHIC_FONT, 15, "bold"),
                     bg=BG, fg="#333", anchor="w", justify="left").pack(anchor="w", pady=(10, 5))
        elif kind == "p":
            _render_text_with_links(inner, value, font=(GOTHIC_FONT, 13),
                                    fg="#333", wraplength=600, pady=(0, 5), bg=BG)
        elif kind == "note":
            f = tk.Frame(inner, bg="#FFF9C4", padx=10, pady=8)
            f.pack(anchor="w", fill="x", pady=5)
            tk.Label(f, text=value, font=(GOTHIC_FONT, 13),
                     bg="#FFF9C4", fg="#555", anchor="w", justify="left",
                     wraplength=580).pack(anchor="w")
        elif kind == "ol":
            for i, item in enumerate(value, 1):
                _render_text_with_links(inner, f"  {i}. {item}",
                                        font=(GOTHIC_FONT, 13), fg="#333",
                                        wraplength=580, pady=(0, 1), bg=BG)
        elif kind == "ul":
            for item in value:
                _render_text_with_links(inner, f"  •  {item}",
                                        font=(GOTHIC_FONT, 13), fg="#333",
                                        wraplength=580, pady=(0, 1), bg=BG)
        elif kind == "code":
            # Bloc de code monospace sur fond gris clair.
            f = tk.Frame(inner, bg="#F5F5F5", padx=10, pady=8)
            f.pack(anchor="w", fill="x", pady=5)
            tk.Label(f, text=value, font=(GOTHIC_FONT, 13),
                     bg="#F5F5F5", fg="#222", anchor="w", justify="left").pack(anchor="w")
        elif kind == "sep":
            # Séparateur horizontal fin.
            tk.Frame(inner, height=1, bg="#DDD").pack(fill="x", pady=15)
        elif kind == "table":
            # value = (headers, rows) — tableau 3 colonnes simple en Grid.
            headers, rows = value
            tbl = tk.Frame(inner, bg=BG)
            tbl.pack(anchor="w", fill="x", pady=5)
            col_widths = (22, 32, 32)
            for col, h in enumerate(headers):
                tk.Label(tbl, text=h, font=(GOTHIC_FONT, 13, "bold"),
                         bg="#E8F5E9", fg="#1B5E20", anchor="w",
                         width=col_widths[col], padx=6, pady=4,
                         borderwidth=1, relief="solid").grid(row=0, column=col, sticky="nsew")
            for r, row in enumerate(rows, start=1):
                bg = "#E8E8E8" if r % 2 == 0 else "#F5F5F5"
                for col, cell in enumerate(row):
                    tk.Label(tbl, text=cell, font=(GOTHIC_FONT, 13),
                             bg=bg, fg="#333", anchor="w", justify="left",
                             width=col_widths[col], padx=6, pady=4,
                             wraplength=col_widths[col] * 7,
                             borderwidth=1, relief="solid").grid(row=r, column=col, sticky="nsew")

    # Bouton fermer en bas
    tk.Button(inner, text="Fermer", command=win.destroy,
              bg="#2E7D32", fg="white", font=(GOTHIC_FONT, 13, "bold"),
              padx=20, pady=5, relief="flat", cursor="hand2").pack(pady=20)


# --------- Interface graphique ---------

# Couleurs du bouton principal
BTN_READY_BG = "#4CAF50"      # vert actif
BTN_READY_HOVER = "#43A047"   # vert hover
BTN_DISABLED_BG = "#BDBDBD"   # gris inactif
BTN_RUNNING_BG = "#FF9800"    # orange pendant l'analyse

# Couleur de fond de l'interface
APP_BG = "#EFEFEF"
GOTHIC_FONT = None  # Résolu au démarrage de l'App (après que Tk soit initialisé)


def _resource_path(filename):
    """Renvoie le chemin absolu vers une ressource, compatible PyInstaller."""
    base = getattr(sys, "_MEIPASS", Path(__file__).parent)
    return str(Path(base) / filename)


class App:
    def __init__(self, root):
        self.root = root
        root.title("Zoquez la mdt bords")
        root.geometry("720x640")
        root.resizable(False, False)
        root.configure(bg=APP_BG)

        # Police gothique : Old English Text MT (macOS/Windows), sinon Century Gothic
        global GOTHIC_FONT
        from tkinter import font as tkfont
        available = tkfont.families()
        if "Old English Text MT" in available:
            GOTHIC_FONT = "Old English Text MT"
        elif "Blackmoor LET" in available:
            GOTHIC_FONT = "Blackmoor LET"
        else:
            GOTHIC_FONT = "Century Gothic"
        root.option_add("*Font", (GOTHIC_FONT, 14))

        # Barre supérieure avec bouton aide
        top_bar = tk.Frame(root, bg=APP_BG)
        top_bar.pack(fill="x", padx=10, pady=(10, 0))
        tk.Button(top_bar, text="❓ Aide", command=lambda: show_help(root),
                  font=(GOTHIC_FONT, 13), padx=12, pady=4, cursor="hand2",
                  relief="flat", bg="#E3F2FD", fg="#1565C0").pack(side="right")

        # En-tête
        header = tk.Frame(root, bg=APP_BG)
        header.pack(fill="x", padx=20, pady=(8, 8))

        title_frame = tk.Frame(header, bg=APP_BG)
        title_frame.pack(side="left", expand=True, fill="x")
        tk.Label(title_frame, text="📧 Analyseur de boite mail en local",
                 font=(GOTHIC_FONT, 18, "bold"), bg=APP_BG, anchor="w").pack(anchor="w")
        tk.Label(title_frame,
                 text="Récupère un fichier Excel avec tous les sites où tu as un compte\nlié à ton adresse gadz.org.\n Tout est traité localement sur ton ordinateur (analyse par mots clés sans IA).\nAucune donnée n'est envoyée sur Internet.",
                 font=(GOTHIC_FONT, 13), fg="#555", bg=APP_BG, anchor="w", justify="left").pack(anchor="w")

        # Variables reliées aux champs (avec trace pour réactivité)
        self.mbox_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.mbox_path.trace_add("write", self._on_field_change)
        self.output_dir.trace_add("write", self._on_field_change)

        # Étape 1 — Télécharger l'export Google Takeout
        f0 = tk.Frame(root, bg=APP_BG)
        f0.pack(pady=(0, 2), fill="x", padx=20)
        tk.Label(f0, text="1. Télécharge ton fichier MBOX (export Gmail via Google Takeout) :",
                 anchor="w", bg=APP_BG).pack(anchor="w")
        link = tk.Label(f0, text="→ takeout.google.com", anchor="w",
                        bg=APP_BG, fg="#1565C0", cursor="hand2",
                        font=(GOTHIC_FONT, 13, "underline"))
        link.pack(anchor="w")
        link.bind("<Button-1>", lambda _: webbrowser.open("https://takeout.google.com"))

        # Sélection du fichier MBOX
        f1 = tk.Frame(root, bg=APP_BG)
        f1.pack(pady=5, fill="x", padx=20)
        tk.Label(f1, text="2. Importe le fichier MBOX ici :", anchor="w", bg=APP_BG).pack(anchor="w")
        f1b = tk.Frame(f1, bg=APP_BG)
        f1b.pack(fill="x")
        tk.Entry(f1b, textvariable=self.mbox_path).pack(side="left", fill="x", expand=True)
        tk.Button(f1b, text="Parcourir...", command=self.choose_mbox).pack(side="right", padx=(5, 0))

        # Sélection du dossier de sortie
        f2 = tk.Frame(root, bg=APP_BG)
        f2.pack(pady=10, fill="x", padx=20)
        tk.Label(f2, text="3. Dossier où enregistrer les résultats (tu peux le changer) :", anchor="w", bg=APP_BG).pack(anchor="w")
        f2b = tk.Frame(f2, bg=APP_BG)
        f2b.pack(fill="x")
        tk.Entry(f2b, textvariable=self.output_dir).pack(side="left", fill="x", expand=True)
        tk.Button(f2b, text="Parcourir...", command=self.choose_output).pack(side="right", padx=(5, 0))

        # Barre de progression
        self.progress = ttk.Progressbar(root, length=580, mode="determinate")
        self.progress.pack(pady=15)

        self.status = tk.Label(root, text="Remplis les deux champs pour lancer l'analyse",
                               fg="#888", bg=APP_BG)
        self.status.pack()

        self.duration_hint = tk.Label(root, text="", fg="#E65100",
                                      font=(GOTHIC_FONT, 13, "italic"), bg=APP_BG)
        self.duration_hint.pack()

        # Ligne de boutons (lancer + télécharger résultat côte à côte)
        btn_row = tk.Frame(root, bg=APP_BG)
        btn_row.pack(pady=10)

        self.btn = tk.Button(btn_row, text="4.  ▶  Lancer l'analyse",
                             command=self.start, bg=BTN_DISABLED_BG, fg="black",
                             font=(GOTHIC_FONT, 15, "bold"), padx=30, pady=10,
                             relief="flat", cursor="arrow",
                             state="disabled",
                             disabledforeground="black")
        self.btn.pack(side="left")

        # Hover effects
        self.btn.bind("<Enter>", self._on_btn_hover)
        self.btn.bind("<Leave>", self._on_btn_leave)

        # Bouton "Télécharger Résultat" — caché au départ, apparaît sur la même ligne en fin d'analyse.
        self.last_result_path = None
        self.open_result_btn = tk.Button(
            btn_row, text="📥  5. Télécharger Résultat",
            command=self._open_last_result,
            bg="#1565C0", fg="black",
            font=(GOTHIC_FONT, 15, "bold"), padx=20, pady=10,
            relief="flat", cursor="hand2",
        )
        # non-packé au démarrage — apparaît après le premier succès.

        tk.Label(root, text="v54me209mdt\nValidez la strass", fg="#AAAAAA",
                 font=(GOTHIC_FONT, 11), bg=APP_BG).pack(side="bottom", pady=(0, 4))

    def _fields_ready(self):
        return bool(self.mbox_path.get().strip() and self.output_dir.get().strip())

    def _on_field_change(self, *args):
        """Appelé dès qu'un champ change. Active/désactive le bouton."""
        if self._fields_ready():
            self.btn.config(state="normal", bg=BTN_READY_BG, fg="black", cursor="hand2")
            self.status.config(text="Prêt à lancer l'analyse", fg="#2E7D32")
        else:
            self.btn.config(state="disabled", bg=BTN_DISABLED_BG, fg="black", cursor="arrow")
            self.status.config(text="Remplissez les deux champs pour lancer l'analyse", fg="#888")

    def _on_btn_hover(self, event):
        if self.btn["state"] == "normal" and self.btn["bg"] == BTN_READY_BG:
            self.btn.config(bg=BTN_READY_HOVER)

    def _on_btn_leave(self, event):
        if self.btn["state"] == "normal" and self.btn["bg"] == BTN_READY_HOVER:
            self.btn.config(bg=BTN_READY_BG)

    def choose_mbox(self):
        path = filedialog.askopenfilename(
            title="Sélectionnez le fichier MBOX",
            filetypes=[("Fichiers MBOX", "*.mbox"), ("Tous les fichiers", "*.*")]
        )
        if path:
            self.mbox_path.set(path)
            self.output_dir.set(str(Path(path).parent / "resultats_mbox"))

    def choose_output(self):
        path = filedialog.askdirectory(title="Choisissez le dossier de sortie")
        if path:
            self.output_dir.set(path)

    def _open_last_result(self):
        if not self.last_result_path or not Path(self.last_result_path).exists():
            messagebox.showerror("Erreur", "Le fichier de résultats est introuvable.")
            return
        if not _open_file_with_default_app(self.last_result_path):
            messagebox.showerror(
                "Erreur",
                f"Impossible d'ouvrir automatiquement :\n{self.last_result_path}\n\n"
                "Ouvrez-le manuellement depuis le Finder/Explorateur."
            )

    def update_progress(self, current, total):
        pct = int(100 * current / total) if total else 0
        self.progress["value"] = pct
        self.status.config(text=f"Analyse en cours... {current:,} / {total:,} messages traités", fg="#E65100")
        self.root.update_idletasks()

    def start(self):
        if not self._fields_ready():
            return
        if not Path(self.mbox_path.get()).exists():
            messagebox.showerror("Erreur", "Le fichier MBOX est introuvable.")
            return

        self.btn.config(state="disabled", text="Analyse en cours...", bg=BTN_RUNNING_BG, fg="black", cursor="arrow")
        self.duration_hint.config(text="⏱️  Selon la taille de votre boîte, l'analyse peut prendre entre 5 et 10 minutes.")
        # Une nouvelle analyse invalide le bouton "Ouvrir" précédent.
        self.open_result_btn.pack_forget()
        self.last_result_path = None

        def run():
            try:
                result = analyse(self.mbox_path.get(), self.output_dir.get(), self.update_progress)
                xlsx_path = str(Path(self.output_dir.get()) / "resultats.xlsx")
                self.last_result_path = xlsx_path
                self.status.config(text="✅ Analyse terminée !", fg="#2E7D32")
                # Affiche le bouton "Ouvrir" côté UI-thread.
                self.root.after(0, lambda: self.open_result_btn.pack(side="left", padx=(12, 0)))
                messagebox.showinfo(
                    "Terminé",
                    f"Analyse réussie !\n\n"
                    f"• {result['analyses']:,} messages analysés\n"
                    f"• {result['ignores']:,} messages ignorés (envoyés/brouillons)\n"
                    f"• {result['total_domaines']} expéditeurs uniques\n"
                    f"• {result['comptes']} comptes probables à traiter\n"
                    f"• {result['newsletters']} newsletters à résilier\n"
                    + (f"• ⚠️  {result['spam']} domaine(s) dans le Spam\n" if result.get('spam') else "")
                    + f"\nFichier Excel généré :\n{xlsx_path}"
                )
            except Exception as e:
                messagebox.showerror("Erreur", f"Une erreur est survenue :\n{e}")
            finally:
                self.btn.config(text="▶  Lancer l'analyse", fg="black")
                self.duration_hint.config(text="")
                self._on_field_change()  # réévalue l'état du bouton

        threading.Thread(target=run, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
